from __future__ import annotations

import csv
import hashlib
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from freight.models import (
    Agent,
    Carrier,
    CarrierService,
    ImportJob,
    Platform,
    PlatformCarrier,
    QuoteChannel,
    RateCard,
    RateRule,
    RateZone,
    SurchargeRule,
    Warehouse,
    WarehouseCarrier,
    WarehousePlatform,
)


SOURCE_SYSTEM = "UBI_TGE_IPEC"
PUBLIC_ZONE_SOURCE = "Maropost/Neto TollIPEC-ShippingZones.csv"
PUBLIC_ZONE_URL = "https://neto.com.au/assets/docs/8007/TollIPEC-ShippingZones.csv"
DEFAULT_ZONE_FILE = r"C:\Users\KenHu\.vscode\CourieDelivery\outputs\ipec_zone_lookup\TollIPEC-ShippingZones.csv"
DEFAULT_UBI_ZIP = r"C:\Users\KenHu\Downloads\ubi_invoices_all 1.zip"
CALCULATOR_KEY = "freight.calculators.ubi_tge_ipec.UbiTgeIpecCalculator"
ORIGINS = ("MEL1", "SYD1")


@dataclass(frozen=True)
class IpecVersion:
    sheet_name: str
    version: str
    effective_from: date
    effective_to: date | None
    source_file: str


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_upper(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value).upper()).strip()


def clean_postcode(value: Any) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+(\.0+)?", text):
        text = text.split(".", 1)[0]
    return text.zfill(4) if text.isdigit() and len(text) < 4 else text


def dec(value: Any, default: str = "0") -> Decimal:
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        text = default
    try:
        return Decimal(text)
    except (InvalidOperation, AttributeError):
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return Decimal(match.group(0)) if match else Decimal(default)


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def zone_prefix(value: Any) -> str:
    text = clean_upper(value).replace("-", " ")
    match = re.match(r"([A-Z]+\d?)", text)
    return match.group(1) if match else text


def infer_state(postcode: str) -> str:
    try:
        number = int(postcode)
    except ValueError:
        return ""
    if 200 <= number <= 299 or 2600 <= number <= 2618 or 2900 <= number <= 2920:
        return "ACT"
    if 1000 <= number <= 2599 or 2619 <= number <= 2899 or 2921 <= number <= 2999:
        return "NSW"
    if 3000 <= number <= 3999 or 8000 <= number <= 8999:
        return "VIC"
    if 4000 <= number <= 4999 or 9000 <= number <= 9999:
        return "QLD"
    if 5000 <= number <= 5999:
        return "SA"
    if 6000 <= number <= 6999:
        return "WA"
    if 7000 <= number <= 7999:
        return "TAS"
    if 800 <= number <= 999:
        return "NT"
    return ""


class Command(BaseCommand):
    help = "Import UBI Team Global Express IPEC rates and fill missing postcode/suburb mapping from public Toll IPEC zones."

    def add_arguments(self, parser):
        parser.add_argument("--zone-file", default=DEFAULT_ZONE_FILE)
        parser.add_argument("--ubi-zip", default=DEFAULT_UBI_ZIP)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--open-all-access", action="store_true")
        parser.add_argument("--apply-billing-overrides", action="store_true", default=True)
        parser.add_argument("--skip-billing-overrides", action="store_false", dest="apply_billing_overrides")
        parser.add_argument("--fuel-rate", default="0.099")
        parser.add_argument("--carrier-name", default="Team Global Express")
        parser.add_argument("--service-code", default="UBI_TGE_IPEC_ROAD")
        parser.add_argument("--service-name", default="UBI TGE IPEC Road")

    def handle(self, *args, **options):
        zone_path = Path(options["zone_file"])
        ubi_zip = Path(options["ubi_zip"])
        if not zone_path.exists():
            raise CommandError(f"IPEC zone file not found: {zone_path}")
        if not ubi_zip.exists():
            raise CommandError(f"UBI zip not found: {ubi_zip}")

        zone_rows = self._parse_public_zone_rows(zone_path)
        versions = self._parse_ipec_rate_versions(ubi_zip)
        if not versions:
            raise CommandError("No IPEC Rate sheets found in UBI zip.")
        billing_overrides, billing_report = self._parse_billing_overrides(ubi_zip, zone_rows) if options["apply_billing_overrides"] else ({}, {})
        merged_zone_rows = self._merge_overrides(zone_rows, billing_overrides)
        report = self._source_report(zone_path, ubi_zip, versions, zone_rows, merged_zone_rows, billing_report, options)

        if options["dry_run"]:
            self._print_report(report, dry_run=True)
            return

        job = ImportJob.objects.create(
            job_type=ImportJob.JobType.RATE_CARD,
            status=ImportJob.Status.RUNNING,
            total_rows=report["total_import_rows"],
            report_json={"source": SOURCE_SYSTEM, "zone_file": str(zone_path), "ubi_zip": str(ubi_zip)},
        )
        try:
            with transaction.atomic():
                import_report = self._import(options, versions, merged_zone_rows, report)
                if options["open_all_access"]:
                    import_report["open_all_access"] = self._open_all_access(import_report["carrier_id"], import_report["service_id"])
        except Exception as exc:  # noqa: BLE001
            job.status = ImportJob.Status.FAILED
            job.error_rows = report["total_import_rows"]
            job.progress = 100
            job.report_json = {**job.report_json, "error": str(exc)}
            job.save(update_fields=["status", "error_rows", "progress", "report_json", "updated_at"])
            raise

        job.status = ImportJob.Status.COMPLETED
        job.success_rows = report["total_import_rows"]
        job.error_rows = 0
        job.progress = 100
        job.report_json = {**job.report_json, **import_report}
        job.save(update_fields=["status", "success_rows", "error_rows", "progress", "report_json", "updated_at"])
        self._print_report(import_report, dry_run=False)

    def _parse_public_zone_rows(self, path: Path) -> list[dict[str, Any]]:
        rows = []
        seen = set()
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            required = {"From Post Code", "Zone Code", "City/Suburb"}
            if not required.issubset(set(reader.fieldnames or [])):
                raise CommandError(f"IPEC zone CSV missing columns: {sorted(required)}")
            for row in reader:
                postcode = clean_postcode(row.get("From Post Code"))
                suburb = clean_upper(row.get("City/Suburb"))
                dest_zone = zone_prefix(row.get("Zone Code"))
                if not postcode or not suburb or not dest_zone:
                    continue
                state = infer_state(postcode)
                key = (postcode, suburb, state, dest_zone)
                if key in seen:
                    continue
                seen.add(key)
                rows.append(
                    {
                        "postcode": postcode,
                        "suburb": suburb,
                        "state": state,
                        "dest_zone": dest_zone,
                        "raw_zone_code": clean(row.get("Zone Code")),
                        "raw_zone_name": clean(row.get("Zone Name")),
                        "source": PUBLIC_ZONE_SOURCE,
                        "source_confidence": "PUBLIC_NETO_REFERENCE",
                    }
                )
        return rows

    def _parse_ipec_rate_versions(self, zip_path: Path) -> dict[IpecVersion, list[dict[str, Any]]]:
        versions: dict[str, tuple[IpecVersion, list[dict[str, Any]]]] = {}
        with ZipFile(zip_path) as archive:
            workbook_names = [name for name in archive.namelist() if name.lower().startswith("ubi_invoices_all/toll/") and name.lower().endswith(".xlsx")]
            for workbook_name in workbook_names:
                workbook = load_workbook(BytesIO(archive.read(workbook_name)), data_only=True, read_only=True)
                for sheet_name in workbook.sheetnames:
                    if not sheet_name.startswith("IPEC Rate "):
                        continue
                    effective_from = self._effective_date(sheet_name)
                    if not effective_from:
                        continue
                    version = IpecVersion(
                        sheet_name=sheet_name,
                        version=f"UBI-IPEC-{effective_from:%Y%m%d}",
                        effective_from=effective_from,
                        effective_to=None,
                        source_file=Path(workbook_name).name,
                    )
                    if sheet_name in versions:
                        continue
                    versions[sheet_name] = (version, self._parse_rate_sheet(workbook[sheet_name], sheet_name))
        ordered = sorted(versions.values(), key=lambda item: item[0].effective_from)
        with_effective_to: dict[IpecVersion, list[dict[str, Any]]] = {}
        for index, (version, rows) in enumerate(ordered):
            next_from = ordered[index + 1][0].effective_from if index + 1 < len(ordered) else None
            effective_to = next_from - timedelta(days=1) if next_from else None
            version = IpecVersion(version.sheet_name, version.version, version.effective_from, effective_to, version.source_file)
            with_effective_to[version] = rows
        return with_effective_to

    def _effective_date(self, sheet_name: str) -> date | None:
        match = re.search(r"(\d{2})\.(\d{2})\.(\d{2,4})", sheet_name)
        if not match:
            return None
        day, month, year = match.groups()
        full_year = int(year) if len(year) == 4 else 2000 + int(year)
        return date(full_year, int(month), int(day))

    def _parse_rate_sheet(self, ws, sheet_name: str) -> list[dict[str, Any]]:
        header_row = None
        headers = []
        for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [clean(value) for value in row]
            if {"service_category", "sce_zone", "dest_zone", "Minimum Charge"}.issubset(set(values)):
                header_row = index
                headers = values
                break
        if not header_row:
            raise CommandError(f"Could not locate IPEC rate header in {sheet_name}")
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            data = {headers[index]: value for index, value in enumerate(row[: len(headers)]) if headers[index]}
            service_category = clean(data.get("service_category"))
            origin = clean_upper(data.get("sce_zone"))
            dest_zone = clean_upper(data.get("dest_zone"))
            if service_category != "Road Express" or origin not in ORIGINS or not dest_zone:
                continue
            rows.append(
                {
                    "origin_zone": origin,
                    "dest_zone": dest_zone,
                    "code": clean(data.get("CODE")),
                    "minimum_charge": dec(data.get("Minimum Charge")),
                    "basic_charge": dec(data.get("BasicChargeAmount")),
                    "freight_charge": dec(data.get("FreightChargeAmount")),
                    "kg_included_in_basic": dec(data.get("KG Included In Basic")),
                    "reciprocal": clean(data.get("Reciprical")),
                    "cubic_conversion": dec(data.get("Cubic Conversion"), "250"),
                    "source_sheet": sheet_name,
                }
            )
        return rows

    def _parse_billing_overrides(self, zip_path: Path, public_rows: list[dict[str, Any]]) -> tuple[dict[tuple[str, str, str], dict[str, Any]], dict[str, Any]]:
        public_exact = {(row["postcode"], row["suburb"], row["state"]): row["dest_zone"] for row in public_rows}
        public_postcode = defaultdict(set)
        for row in public_rows:
            public_postcode[(row["postcode"], row["state"])].add(row["dest_zone"])

        observed: dict[tuple[str, str, str], Counter] = defaultdict(Counter)
        stats = Counter()
        with ZipFile(zip_path) as archive:
            workbook_names = [name for name in archive.namelist() if name.lower().startswith("ubi_invoices_all/toll/") and name.lower().endswith(".xlsx")]
            for workbook_name in workbook_names:
                workbook = load_workbook(BytesIO(archive.read(workbook_name)), data_only=True, read_only=True)
                if "Billing" not in workbook.sheetnames:
                    continue
                ws = workbook["Billing"]
                headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                header_map = {header: index for index, header in enumerate(headers)}
                required = {"City/Suburb", "State", "Postcode", "To Zone"}
                if not required.issubset(header_map):
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    postcode = clean_postcode(row[header_map["Postcode"]])
                    suburb = clean_upper(row[header_map["City/Suburb"]])
                    state = clean_upper(row[header_map["State"]])
                    zone = zone_prefix(row[header_map["To Zone"]])
                    if not postcode or not suburb or not state or not zone:
                        continue
                    stats["billing_rows"] += 1
                    key = (postcode, suburb, state)
                    observed[key][zone] += 1
                    if public_exact.get(key) == zone:
                        stats["public_exact_match"] += 1
                    elif key in public_exact:
                        stats["public_exact_mismatch"] += 1
                    elif public_postcode.get((postcode, state)) == {zone}:
                        stats["public_postcode_unique_match"] += 1
                    elif (postcode, state) not in public_postcode:
                        stats["public_missing_postcode_state"] += 1

        overrides = {}
        for key, counts in observed.items():
            zone, count = counts.most_common(1)[0]
            public_zone = public_exact.get(key)
            if public_zone == zone:
                continue
            if public_zone and count < 2:
                continue
            postcode, suburb, state = key
            overrides[key] = {
                "postcode": postcode,
                "suburb": suburb,
                "state": state,
                "dest_zone": zone,
                "observed_count": count,
                "observed_zones": dict(counts),
                "public_zone": public_zone or "",
                "source": "UBI Toll IPEC billing",
                "source_confidence": "UBI_BILLING_OBSERVED_OVERRIDE",
            }
        return overrides, {"stats": dict(stats), "override_rows": len(overrides)}

    def _merge_overrides(
        self,
        public_rows: list[dict[str, Any]],
        overrides: dict[tuple[str, str, str], dict[str, Any]],
    ) -> list[dict[str, Any]]:
        rows = list(public_rows)
        for row in overrides.values():
            rows.append(row)
        return rows

    def _source_report(
        self,
        zone_path: Path,
        ubi_zip: Path,
        versions: dict[IpecVersion, list[dict[str, Any]]],
        public_rows: list[dict[str, Any]],
        merged_rows: list[dict[str, Any]],
        billing_report: dict[str, Any],
        options: dict[str, Any],
    ) -> dict[str, Any]:
        version_rows = {version.version: len(rows) for version, rows in versions.items()}
        zones_by_version = {version.version: sorted({row["dest_zone"] for row in rows}) for version, rows in versions.items()}
        mapping_zones = sorted({row["dest_zone"] for row in merged_rows})
        missing_by_version = {version: sorted(set(zones) - set(mapping_zones)) for version, zones in zones_by_version.items()}
        return {
            "source": SOURCE_SYSTEM,
            "public_zone_url": PUBLIC_ZONE_URL,
            "zone_file": str(zone_path),
            "zone_file_sha256": file_sha256(zone_path),
            "ubi_zip": str(ubi_zip),
            "ubi_zip_sha256": file_sha256(ubi_zip),
            "versions": [
                {
                    "version": version.version,
                    "sheet_name": version.sheet_name,
                    "source_file": version.source_file,
                    "effective_from": str(version.effective_from),
                    "effective_to": str(version.effective_to) if version.effective_to else None,
                    "rate_rows": len(rows),
                }
                for version, rows in versions.items()
            ],
            "rate_rows_by_version": version_rows,
            "public_zone_rows": len(public_rows),
            "merged_zone_rows": len(merged_rows),
            "billing_override_rows": max(0, len(merged_rows) - len(public_rows)),
            "mapping_zones": len(mapping_zones),
            "missing_zone_mappings": missing_by_version,
            "billing_report": billing_report,
            "fuel_rate": str(dec(options["fuel_rate"])),
            "total_import_rows": sum(len(rows) for rows in versions.values()) + len(merged_rows) * len(versions),
        }

    def _import(
        self,
        options: dict[str, Any],
        versions: dict[IpecVersion, list[dict[str, Any]]],
        zone_rows: list[dict[str, Any]],
        source_report: dict[str, Any],
    ) -> dict[str, Any]:
        agent = self._agent()
        carrier = self._carrier(options["carrier_name"])
        service = self._service(carrier, options["service_code"], options["service_name"])
        report = {**source_report, "agent_id": agent.id, "carrier_id": carrier.id, "service_id": service.id, "rate_cards": 0, "rate_rules": 0, "rate_zones": 0, "surcharge_rules": 0, "quote_channels": 0}
        for version, rows in versions.items():
            by_origin = {origin: [row for row in rows if row["origin_zone"] == origin] for origin in ORIGINS}
            for origin, origin_rows in by_origin.items():
                if not origin_rows:
                    continue
                card = self._rate_card(carrier, service, version, origin, origin_rows, source_report)
                card.rules.all().delete()
                card.zones.all().delete()
                card.surcharge_rules.all().delete()
                rules = self._rate_rules(card, service, origin_rows)
                zones = self._rate_zones(card, origin, zone_rows, {row["dest_zone"] for row in origin_rows})
                surcharges = self._surcharge_rules(card, options["fuel_rate"])
                RateRule.objects.bulk_create(rules, batch_size=1000)
                RateZone.objects.bulk_create(zones, batch_size=1000)
                SurchargeRule.objects.bulk_create(surcharges, batch_size=100)
                self._quote_channel(agent, carrier, service, card, version, origin)
                report["rate_cards"] += 1
                report["rate_rules"] += len(rules)
                report["rate_zones"] += len(zones)
                report["surcharge_rules"] += len(surcharges)
                report["quote_channels"] += 1
        return report

    def _agent(self) -> Agent:
        agent, _ = Agent.objects.update_or_create(
            code="ubi",
            defaults={
                "name": "UBI",
                "agent_type": Agent.AgentType.LSP,
                "active": True,
                "supports_api": True,
                "maintains_rate_cards": True,
                "source_system": SOURCE_SYSTEM,
            },
        )
        return agent

    def _carrier(self, carrier_name: str) -> Carrier:
        carrier = Carrier.objects.filter(name__iexact=carrier_name).order_by("id").first()
        if not carrier:
            carrier = Carrier.objects.create(
                code=self._next_carrier_code(),
                name=carrier_name,
                carrier_type=Carrier.CarrierType.HYBRID,
                active=True,
                support_api=True,
                lsp_agent_code="ubi",
                lsp_channel_code="toll",
                source_system=SOURCE_SYSTEM,
                source_external_id="UBI.AU2AU.IPEC",
            )
        payload = dict(carrier.source_payload_json or {})
        payload["ubi_ipec"] = {"source": SOURCE_SYSTEM, "lsp_carrier_code": "UBI.AU2AU.IPEC", "last_imported_at": timezone.now().isoformat()}
        carrier.name = carrier_name
        carrier.carrier_type = Carrier.CarrierType.HYBRID
        carrier.active = True
        carrier.support_api = True
        carrier.lsp_agent_code = "ubi"
        carrier.lsp_channel_code = "toll"
        carrier.source_payload_json = payload
        carrier.save(update_fields=["name", "carrier_type", "active", "support_api", "lsp_agent_code", "lsp_channel_code", "source_payload_json", "updated_at"])
        return carrier

    def _next_carrier_code(self) -> str:
        max_number = 0
        for code in Carrier.objects.filter(code__startswith="CAR").values_list("code", flat=True):
            match = re.fullmatch(r"CAR(\d{6})", code or "")
            if match:
                max_number = max(max_number, int(match.group(1)))
        number = max_number + 1
        while True:
            code = f"CAR{number:06d}"
            if not Carrier.objects.filter(code=code).exists():
                return code
            number += 1

    def _service(self, carrier: Carrier, code: str, name: str) -> CarrierService:
        service, _ = CarrierService.objects.update_or_create(
            carrier=carrier,
            code=code,
            defaults={"name": name, "service_level": "IPEC Road", "active": True},
        )
        return service

    def _rate_card(
        self,
        carrier: Carrier,
        service: CarrierService,
        version: IpecVersion,
        origin: str,
        rows: list[dict[str, Any]],
        source_report: dict[str, Any],
    ) -> RateCard:
        metadata = {
            "source": SOURCE_SYSTEM,
            "agent_code": "ubi",
            "source_zone_file": source_report["zone_file"],
            "source_zone_file_sha256": source_report["zone_file_sha256"],
            "source_public_zone_url": PUBLIC_ZONE_URL,
            "source_ubi_zip": source_report["ubi_zip"],
            "source_ubi_zip_sha256": source_report["ubi_zip_sha256"],
            "source_workbook": version.source_file,
            "source_sheet": version.sheet_name,
            "origin_zone": origin,
            "calculator_key": CALCULATOR_KEY,
            "pricing_formula": "max(minimum_charge, basic_charge + freight_charge_per_kg * max(0, chargeable_kg - kg_included_in_basic))",
            "mapping_source_confidence": "PUBLIC_NETO_REFERENCE_WITH_UBI_BILLING_OVERRIDES",
            "destination_zones": sorted({row["dest_zone"] for row in rows}),
        }
        card, _ = RateCard.objects.update_or_create(
            legacy_source_object=f"{SOURCE_SYSTEM}:{version.version}:{origin}",
            version=f"{version.version}-{origin}",
            defaults={
                "carrier": carrier,
                "service": service,
                "name": f"UBI TGE IPEC Road {origin} {version.effective_from:%Y-%m-%d}",
                "version_label": version.version,
                "status": RateCard.Status.ACTIVE,
                "effective_from": version.effective_from,
                "effective_to": version.effective_to,
                "is_active": True,
                "priority": 90 if origin == "MEL1" else 91,
                "currency": "AUD",
                "tax_mode": RateCard.TaxMode.EX_GST,
                "gst_rate": Decimal("0.10"),
                "cubic_factor": Decimal("250"),
                "metadata_json": metadata,
            },
        )
        return card

    def _rate_rules(self, card: RateCard, service: CarrierService, rows: list[dict[str, Any]]) -> list[RateRule]:
        rules = []
        for index, row in enumerate(rows, start=1):
            rules.append(
                RateRule(
                    rate_card=card,
                    service=service,
                    from_zone=row["origin_zone"],
                    to_zone=row["dest_zone"],
                    weight_min_kg=Decimal("0"),
                    weight_max_kg=None,
                    basic_charge=row["basic_charge"],
                    per_kg=row["freight_charge"],
                    minimum_charge=row["minimum_charge"],
                    rule_type=RateRule.RuleType.LINEHAUL,
                    priority=index,
                    raw_payload={
                        "source": SOURCE_SYSTEM,
                        "source_sheet": row["source_sheet"],
                        "code": row["code"],
                        "service_category": "Road Express",
                        "kg_included_in_basic": str(row["kg_included_in_basic"]),
                        "reciprocal": row["reciprocal"],
                        "cubic_conversion": str(row["cubic_conversion"]),
                    },
                )
            )
        return rules

    def _rate_zones(
        self,
        card: RateCard,
        origin: str,
        rows: list[dict[str, Any]],
        allowed_zones: set[str],
    ) -> list[RateZone]:
        zone_objects = []
        for row in rows:
            if row["dest_zone"] not in allowed_zones:
                continue
            zone_objects.append(
                RateZone(
                    rate_card=card,
                    origin_zone=origin,
                    dest_zone=row["dest_zone"],
                    state=row["state"],
                    suburb=row["suburb"],
                    postcode=row["postcode"],
                    deliverable=True,
                    raw_payload={
                        "source": row["source"],
                        "source_url": PUBLIC_ZONE_URL if row["source"] == PUBLIC_ZONE_SOURCE else "",
                        "raw_zone_code": row.get("raw_zone_code", ""),
                        "raw_zone_name": row.get("raw_zone_name", ""),
                        "source_confidence": row["source_confidence"],
                        "mapping_precedence": row["source_confidence"],
                        "observed_count": row.get("observed_count"),
                        "observed_zones": row.get("observed_zones"),
                        "public_zone": row.get("public_zone", ""),
                    },
                )
            )
        return zone_objects

    def _surcharge_rules(self, card: RateCard, fuel_rate: str) -> list[SurchargeRule]:
        return [
            SurchargeRule(
                carrier=card.carrier,
                rate_card=card,
                code="FS",
                rule_name="UBI TGE IPEC fuel levy",
                ratio=dec(fuel_rate),
                fee_amount=Decimal("0"),
                match_dimension=SurchargeRule.MatchDimension.ALWAYS,
                priority=1,
                active=True,
                raw_payload={
                    "source": SOURCE_SYSTEM,
                    "note": "Default imported from billing distribution review; edit in Pricing -> Surcharges when UBI publishes a period-specific fuel rate.",
                },
            )
        ]

    def _quote_channel(
        self,
        agent: Agent,
        carrier: Carrier,
        service: CarrierService,
        card: RateCard,
        version: IpecVersion,
        origin: str,
    ) -> QuoteChannel:
        code = f"ubi_tge_ipec_{origin.lower()}"
        channel, _ = QuoteChannel.objects.update_or_create(
            code=code,
            defaults={
                "name": f"UBI TGE IPEC {origin}",
                "carrier": carrier,
                "service": service,
                "provider_type": QuoteChannel.ProviderType.TABLE,
                "calculator_key": CALCULATOR_KEY,
                "quote_source": SOURCE_SYSTEM,
                "enabled": True,
                "priority": 90 if origin == "MEL1" else 91,
                "rate_card": card,
                "agent": agent,
                "config_json": {"source": SOURCE_SYSTEM, "agent_code": "ubi", "origin_zone": origin, "lsp_carrier_code": "UBI.AU2AU.IPEC"},
                "valid_from": None,
                "valid_to": None,
            },
        )
        return channel

    def _open_all_access(self, carrier_id: int, service_id: int) -> dict[str, Any]:
        platforms = list(Platform.objects.filter(active=True))
        warehouses = list(Warehouse.objects.filter(active=True))
        carrier = Carrier.objects.get(id=carrier_id)
        service = CarrierService.objects.get(id=service_id)
        report = {"platforms": len(platforms), "warehouses": len(warehouses), "warehouse_platform": 0, "platform_carrier": 0, "warehouse_carrier": 0}
        for warehouse in warehouses:
            for platform in platforms:
                _, created = WarehousePlatform.objects.get_or_create(
                    warehouse=warehouse,
                    platform=platform,
                    defaults={"enabled": True, "priority": 100},
                )
                if created:
                    report["warehouse_platform"] += 1
            _, created = WarehouseCarrier.objects.update_or_create(
                warehouse=warehouse,
                carrier=carrier,
                service=service,
                defaults={"enabled": True, "origin_zone": self._warehouse_origin(warehouse)},
            )
            if created:
                report["warehouse_carrier"] += 1
        for platform in platforms:
            _, created = PlatformCarrier.objects.update_or_create(
                platform=platform,
                carrier=carrier,
                service=service,
                defaults={"enabled": True, "priority": 90, "quote_source": SOURCE_SYSTEM},
            )
            if created:
                report["platform_carrier"] += 1
        return report

    def _warehouse_origin(self, warehouse: Warehouse) -> str:
        text = f"{warehouse.default_origin_zone} {warehouse.code} {warehouse.name} {warehouse.state} {warehouse.region}".upper()
        if "SYD" in text or "NSW" in text:
            return "SYD1"
        if "MEL" in text or "VIC" in text:
            return "MEL1"
        return ""

    def _print_report(self, report: dict[str, Any], *, dry_run: bool) -> None:
        prefix = "UBI IPEC dry-run" if dry_run else "UBI IPEC import completed"
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix}: rate_cards={report.get('rate_cards', 'n/a')}, "
                f"rate_rules={report.get('rate_rules', report.get('rate_rows_by_version'))}, "
                f"zone_rows={report.get('merged_zone_rows')}, "
                f"billing_overrides={report.get('billing_override_rows')}, "
                f"missing_zone_mappings={report.get('missing_zone_mappings')}"
            )
        )
        if report.get("open_all_access"):
            self.stdout.write(f"open_all_access={report['open_all_access']}")
