from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from freight.models import (
    Carrier,
    CarrierService,
    ImportJob,
    Platform,
    PlatformCarrier,
    QuoteChannel,
    RateCard,
    RateRule,
    RateZone,
    SurchargeRule,
    Warehouse,
    WarehouseCarrier,
    WarehousePlatform,
)
from freight.quote_engine import json_safe


SOURCE_SYSTEM = "DirectFreightExpressProposal"
DEFAULT_RATE_FILE = r"C:\Users\KenHu\Downloads\Direct Feight Express Rates Proposal EX SYD Ex MEL Feb 2025.xlsx"
DEFAULT_ZONE_FILE = r"C:\Users\KenHu\Downloads\Zone List - postcodes 1.csv"
CALCULATOR_KEY = "freight.calculators.direct_freight_express_2025.DirectFreightExpress2025Calculator"


@dataclass(frozen=True)
class DfeSpec:
    key: str
    origin_zone: str
    sheet_name: str
    rate_card_name: str
    version: str
    kilo_service_code: str
    kilo_service_name: str
    pallet_service_code: str
    pallet_service_name: str
    kilo_channel_code: str
    kilo_channel_name: str
    pallet_channel_code: str
    pallet_channel_name: str
    effective_from: date
    priority: int


SPECS = (
    DfeSpec(
        key="dfe_ex_mel_2025",
        origin_zone="MELB",
        sheet_name="Rate EX Mel",
        rate_card_name="DFE EX MEL Feb 2025",
        version="DFE-EX-MEL-FEB-2025",
        kilo_service_code="DFE_KILO_EX_MEL_2025",
        kilo_service_name="DFE KILO EX MEL 2025",
        pallet_service_code="DFE_PALLET_EX_MEL_2025",
        pallet_service_name="DFE PALLET EX MEL 2025",
        kilo_channel_code="dfe_ex_mel_2025",
        kilo_channel_name="DFE EX MEL Feb 2025",
        pallet_channel_code="dfe_pallet_ex_mel_2025",
        pallet_channel_name="DFE PALLET EX MEL Feb 2025",
        effective_from=date(2025, 2, 1),
        priority=70,
    ),
    DfeSpec(
        key="dfe_ex_syd_2025",
        origin_zone="SYDN",
        sheet_name="Ex SYD",
        rate_card_name="DFE EX SYD Feb 2025",
        version="DFE-EX-SYD-FEB-2025",
        kilo_service_code="DFE_KILO_EX_SYD_2025",
        kilo_service_name="DFE KILO EX SYD 2025",
        pallet_service_code="DFE_PALLET_EX_SYD_2025",
        pallet_service_name="DFE PALLET EX SYD 2025",
        kilo_channel_code="dfe_ex_syd_2025",
        kilo_channel_name="DFE EX SYD Feb 2025",
        pallet_channel_code="dfe_pallet_ex_syd_2025",
        pallet_channel_name="DFE PALLET EX SYD Feb 2025",
        effective_from=date(2025, 2, 1),
        priority=71,
    ),
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_upper(value: Any) -> str:
    return clean(value).upper()


def clean_postcode(value: Any) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+(\.0+)?", text):
        text = text.split(".", 1)[0]
    return text.zfill(4) if text.isdigit() and len(text) < 4 else text


def dec(value: Any, default: str = "0") -> Decimal:
    text = clean(value).replace("$", "").replace(",", "")
    if not text or text.upper() in {"POA", "P.O.A", "P.O.A."}:
        text = default
    try:
        return Decimal(text)
    except (InvalidOperation, AttributeError):
        return Decimal(default)


def money_text(value: Decimal) -> str:
    return f"{value:.4f}"


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class Command(BaseCommand):
    help = "Import Direct Freight Express Feb 2025 proposal rates and postcode zones."

    def add_arguments(self, parser):
        parser.add_argument("--rate-file", default=DEFAULT_RATE_FILE)
        parser.add_argument("--zone-file", default=DEFAULT_ZONE_FILE)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--configure-defaults", action="store_true")
        parser.add_argument("--carrier-code", default="454")
        parser.add_argument("--carrier-name", default="Direct Freight Express")
        parser.add_argument("--default-platform-code", default="PI2022080502320043121506")
        parser.add_argument("--default-warehouse-code", default="BG01")

    def handle(self, *args, **options):
        rate_path = Path(options["rate_file"])
        zone_path = Path(options["zone_file"])
        if not rate_path.exists():
            raise CommandError(f"Rate proposal file not found: {rate_path}")
        if not zone_path.exists():
            raise CommandError(f"Zone file not found: {zone_path}")

        rate_book = self._load_workbook_any(rate_path)
        zone_book = self._load_workbook_any(zone_path)
        parsed_rates = {spec.key: self._parse_rate_sheet(rate_book, spec) for spec in SPECS}
        zone_rows = self._parse_zone_rows(zone_book)
        destination_surcharges = self._parse_destination_surcharges(rate_book)
        service_charges = self._parse_service_charge_reference(rate_book)
        report = self._source_report(rate_path, zone_path, parsed_rates, zone_rows, destination_surcharges, service_charges)

        if options["dry_run"]:
            self._print_report(report, dry_run=True)
            return

        total_rows = report["total_import_rows"]
        job = ImportJob.objects.create(
            job_type=ImportJob.JobType.RATE_CARD,
            status=ImportJob.Status.RUNNING,
            total_rows=total_rows,
            report_json={"source": SOURCE_SYSTEM, "rate_file": str(rate_path), "zone_file": str(zone_path)},
        )
        try:
            with transaction.atomic():
                import_report = self._import(
                    options,
                    parsed_rates,
                    zone_rows,
                    destination_surcharges,
                    service_charges,
                    report,
                )
                if options["configure_defaults"]:
                    import_report["default_configuration"] = self._configure_defaults(
                        options["default_platform_code"],
                        options["default_warehouse_code"],
                    )
        except Exception as exc:  # noqa: BLE001
            job.status = ImportJob.Status.FAILED
            job.error_rows = total_rows
            job.progress = 100
            job.report_json = {**job.report_json, "error": str(exc)}
            job.save(update_fields=["status", "error_rows", "progress", "report_json", "updated_at"])
            raise

        job.status = ImportJob.Status.COMPLETED
        job.success_rows = total_rows
        job.error_rows = 0
        job.progress = 100
        job.report_json = {**job.report_json, **import_report}
        job.save(update_fields=["status", "success_rows", "error_rows", "progress", "report_json", "updated_at"])
        self._print_report(import_report, dry_run=False)

    def _load_workbook_any(self, path: Path):
        data = path.read_bytes()
        if data.startswith(b"PK"):
            return load_workbook(BytesIO(data), data_only=True, read_only=True)
        return load_workbook(path, data_only=True, read_only=True)

    def _parse_rate_sheet(self, workbook, spec: DfeSpec) -> list[dict[str, Any]]:
        if spec.sheet_name not in workbook.sheetnames:
            raise CommandError(f"Missing DFE rate sheet: {spec.sheet_name}")
        ws = workbook[spec.sheet_name]
        header_row, headers = self._find_header(ws, ("From", "To", "Basic Charge", "Rate", "Type", "Min Charge"))
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            data = {headers[index]: value for index, value in enumerate(row[: len(headers)]) if headers[index]}
            if not clean(data.get("From")) and not clean(data.get("To")):
                continue
            rate_type = clean_upper(data.get("Type"))
            if rate_type not in {"KILO", "PALLET"}:
                continue
            rows.append(
                {
                    "from_zone": clean_upper(data.get("From")),
                    "to_zone": clean_upper(data.get("To")),
                    "suburb": clean_upper(data.get("Suburb")),
                    "basic_charge": dec(data.get("Basic Charge")),
                    "per_kg": dec(data.get("Rate")),
                    "break_value": dec(data.get("Break")),
                    "fuel_flag": clean(data.get("F/S")),
                    "upto_kg": dec(data.get("Upto KG")),
                    "rate_type": rate_type,
                    "minimum_charge": dec(data.get("Min Charge")),
                    "cubic_factor": dec(data.get("Cubic Conv"), "250"),
                    "source_row": json_safe(data),
                }
            )
        return rows

    def _parse_zone_rows(self, workbook) -> list[dict[str, Any]]:
        ws = workbook.active
        header_row, headers = self._find_header(
            ws,
            ("PostcodeId", "Postcode", "Postcodename", "State", "CarrierZone", "DropCode", "SortCode"),
        )
        rows = []
        seen: set[tuple[str, str, str, str]] = set()
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            data = {headers[index]: value for index, value in enumerate(row[: len(headers)]) if headers[index]}
            postcode = clean_postcode(data.get("Postcode"))
            suburb = clean_upper(data.get("Postcodename"))
            state = clean_upper(data.get("State"))
            carrier_zone = clean_upper(data.get("CarrierZone"))
            if not postcode or not suburb or not state or not carrier_zone:
                continue
            key = (postcode, suburb, state, carrier_zone)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "postcode_id": clean(data.get("PostcodeId")),
                    "postcode": postcode,
                    "suburb": suburb,
                    "state": state,
                    "carrier_zone": carrier_zone,
                    "drop_code": clean(data.get("DropCode")),
                    "sort_code": clean(data.get("SortCode")),
                    "source_row": json_safe(data),
                }
            )
        return rows

    def _parse_destination_surcharges(self, workbook) -> list[dict[str, Any]]:
        if "Surcharge " not in workbook.sheetnames:
            raise CommandError("Missing DFE surcharge sheet: Surcharge ")
        ws = workbook["Surcharge "]
        start_row = None
        for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [clean_upper(value) for value in row[:3]]
            if values[:3] == ["SUBURB", "POSTCODE", "AMOUNT"]:
                start_row = index + 1
                break
        if not start_row:
            raise CommandError("Could not locate DFE destination surcharge table.")
        rows = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            suburb = clean_upper(row[0] if len(row) > 0 else "")
            postcode = clean_postcode(row[1] if len(row) > 1 else "")
            amount = dec(row[2] if len(row) > 2 else None)
            if not suburb or not postcode:
                continue
            rows.append({"suburb": suburb, "postcode": postcode, "amount": amount})
        return rows

    def _parse_service_charge_reference(self, workbook) -> list[dict[str, Any]]:
        ws = workbook["Surcharge "]
        rows = []
        current_section = ""
        for index, row in enumerate(ws.iter_rows(min_row=1, max_row=53, values_only=True), start=1):
            first = clean(row[0] if len(row) > 0 else "")
            second = clean(row[1] if len(row) > 1 else "")
            third = clean(row[2] if len(row) > 2 else "")
            if not first and not second:
                continue
            upper_first = first.upper()
            if second == "" and upper_first not in {"CHARGE CODE", "DESTINATION SURCHARGE"}:
                current_section = first
                continue
            if upper_first == "CHARGE CODE":
                continue
            rows.append(
                {
                    "section": current_section,
                    "charge_code": first,
                    "description": second,
                    "fuel_levy_apply": third,
                    "source_row": index,
                }
            )
        return rows

    def _find_header(self, ws, required: tuple[str, ...]) -> tuple[int, list[str]]:
        required_upper = {item.upper() for item in required}
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            headers = [clean(value) for value in row]
            header_upper = {header.upper() for header in headers if header}
            if required_upper.issubset(header_upper):
                return row_index, headers
        raise CommandError(f"Could not find header row in sheet {ws.title}: {required}")

    def _source_report(
        self,
        rate_path: Path,
        zone_path: Path,
        parsed_rates: dict[str, list[dict[str, Any]]],
        zone_rows: list[dict[str, Any]],
        destination_surcharges: list[dict[str, Any]],
        service_charges: list[dict[str, Any]],
    ) -> dict[str, Any]:
        rate_zones = {
            key: sorted({row["to_zone"] for row in rows if row["to_zone"]})
            for key, rows in parsed_rates.items()
        }
        zone_file_zones = sorted({row["carrier_zone"] for row in zone_rows})
        missing_by_spec = {
            key: sorted(set(zones) - set(zone_file_zones))
            for key, zones in rate_zones.items()
        }
        extra_zone_values = sorted(set(zone_file_zones) - set().union(*(set(zones) for zones in rate_zones.values())))
        return {
            "source": SOURCE_SYSTEM,
            "rate_file": str(rate_path),
            "rate_file_sha256": file_sha256(rate_path),
            "zone_file": str(zone_path),
            "zone_file_sha256": file_sha256(zone_path),
            "rate_rows_by_spec": {key: len(rows) for key, rows in parsed_rates.items()},
            "kilo_rows_by_spec": {key: sum(1 for row in rows if row["rate_type"] == "KILO") for key, rows in parsed_rates.items()},
            "pallet_rows_by_spec": {key: sum(1 for row in rows if row["rate_type"] == "PALLET") for key, rows in parsed_rates.items()},
            "rate_zones_by_spec": {key: len(zones) for key, zones in rate_zones.items()},
            "zone_rows": len(zone_rows),
            "zone_file_zones": len(zone_file_zones),
            "destination_surcharge_rows": len(destination_surcharges),
            "service_charge_reference_rows": len(service_charges),
            "missing_zone_mappings": missing_by_spec,
            "extra_zone_values": extra_zone_values,
            "total_import_rows": sum(len(rows) for rows in parsed_rates.values())
            + len(zone_rows) * len(SPECS)
            + len(destination_surcharges) * len(SPECS)
            + len(SPECS),
        }

    def _import(
        self,
        options: dict[str, Any],
        parsed_rates: dict[str, list[dict[str, Any]]],
        zone_rows: list[dict[str, Any]],
        destination_surcharges: list[dict[str, Any]],
        service_charges: list[dict[str, Any]],
        source_report: dict[str, Any],
    ) -> dict[str, Any]:
        carrier = self._carrier(options["carrier_code"], options["carrier_name"])
        report = {
            **source_report,
            "carrier_id": carrier.id,
            "carrier_code": carrier.code,
            "carrier_name": carrier.name,
            "rate_cards": 0,
            "rate_zones": 0,
            "rate_rules": 0,
            "surcharge_rules": 0,
            "quote_channels": 0,
            "services": [],
        }
        for spec in SPECS:
            kilo_service = self._service(carrier, spec.kilo_service_code, spec.kilo_service_name, "KILO", active=True)
            pallet_service = self._service(carrier, spec.pallet_service_code, spec.pallet_service_name, "PALLET", active=False)
            report["services"].extend([kilo_service.code, pallet_service.code])
            card = self._rate_card(carrier, kilo_service, spec, parsed_rates[spec.key], service_charges, source_report)
            card.rules.all().delete()
            card.zones.all().delete()
            card.surcharge_rules.all().delete()

            rate_rules = self._rate_rules(card, spec, parsed_rates[spec.key], kilo_service, pallet_service)
            zones = self._zones(card, spec, zone_rows, {row["to_zone"] for row in parsed_rates[spec.key]})
            surcharge_rules = self._surcharge_rules(card, destination_surcharges)
            RateRule.objects.bulk_create(rate_rules, batch_size=1000)
            RateZone.objects.bulk_create(zones, batch_size=1000)
            SurchargeRule.objects.bulk_create(surcharge_rules, batch_size=1000)
            self._quote_channel(carrier, kilo_service, card, spec, enabled=True, pallet=False)
            self._quote_channel(carrier, pallet_service, card, spec, enabled=False, pallet=True)

            report["rate_cards"] += 1
            report["rate_zones"] += len(zones)
            report["rate_rules"] += len(rate_rules)
            report["surcharge_rules"] += len(surcharge_rules)
            report["quote_channels"] += 2
        return report

    def _carrier(self, preferred_code: str, carrier_name: str) -> Carrier:
        carrier = Carrier.objects.filter(code=preferred_code).first()
        if not carrier:
            carrier = (
                Carrier.objects.filter(name__icontains="Direct Freight").order_by("id").first()
                or Carrier.objects.filter(name__icontains="Direct").order_by("id").first()
            )
        if not carrier:
            carrier = Carrier.objects.create(
                code=preferred_code,
                name=carrier_name,
                carrier_type=Carrier.CarrierType.TABLE,
                active=True,
                support_api=False,
                source_system=SOURCE_SYSTEM,
            )
        update_fields = []
        original_name = carrier.name
        if carrier.name != carrier_name:
            carrier.name = carrier_name
            update_fields.append("name")
        if carrier.carrier_type == Carrier.CarrierType.API:
            carrier.carrier_type = Carrier.CarrierType.HYBRID
            update_fields.append("carrier_type")
        payload = dict(carrier.source_payload_json or {})
        payload["dfe_rate_proposal"] = {
            "source": SOURCE_SYSTEM,
            "last_imported_at": timezone.now().isoformat(),
            "previous_carrier_name": original_name,
        }
        carrier.source_payload_json = payload
        update_fields.append("source_payload_json")
        if update_fields:
            carrier.save(update_fields=[*set(update_fields), "updated_at"])
        return carrier

    def _service(self, carrier: Carrier, code: str, name: str, service_level: str, *, active: bool) -> CarrierService:
        service, _ = CarrierService.objects.update_or_create(
            carrier=carrier,
            code=code,
            defaults={"name": name, "service_level": service_level, "active": active},
        )
        return service

    def _rate_card(
        self,
        carrier: Carrier,
        service: CarrierService,
        spec: DfeSpec,
        rows: list[dict[str, Any]],
        service_charges: list[dict[str, Any]],
        source_report: dict[str, Any],
    ) -> RateCard:
        cubic_factors = sorted({str(row["cubic_factor"]) for row in rows if row["cubic_factor"]})
        metadata = {
            "source": SOURCE_SYSTEM,
            "source_rate_file": source_report["rate_file"],
            "source_zone_file": source_report["zone_file"],
            "source_rate_file_sha256": source_report["rate_file_sha256"],
            "source_zone_file_sha256": source_report["zone_file_sha256"],
            "source_sheet": spec.sheet_name,
            "origin_zone": spec.origin_zone,
            "calculator_key": CALCULATOR_KEY,
            "fuel_rate": "0.196",
            "profile_restrictions": {
                "max_item_kg": "30",
                "max_longest_cm": "120",
                "two_sides_over_cm": "70",
                "mode": "strict_not_available",
            },
            "service_charge_reference": service_charges,
            "rate_rows": len(rows),
            "kilo_rows": sum(1 for row in rows if row["rate_type"] == "KILO"),
            "pallet_rows": sum(1 for row in rows if row["rate_type"] == "PALLET"),
        }
        card, _ = RateCard.objects.update_or_create(
            legacy_source_object=f"{SOURCE_SYSTEM}:{spec.version}",
            version=spec.version,
            defaults={
                "carrier": carrier,
                "service": service,
                "name": spec.rate_card_name,
                "version_label": spec.version,
                "status": RateCard.Status.ACTIVE,
                "effective_from": spec.effective_from,
                "effective_to": None,
                "is_active": True,
                "priority": spec.priority,
                "currency": "AUD",
                "tax_mode": RateCard.TaxMode.EX_GST,
                "gst_rate": Decimal("0.10"),
                "cubic_factor": dec(cubic_factors[0] if cubic_factors else "250", "250"),
                "metadata_json": metadata,
            },
        )
        return card

    def _rate_rules(
        self,
        card: RateCard,
        spec: DfeSpec,
        rows: list[dict[str, Any]],
        kilo_service: CarrierService,
        pallet_service: CarrierService,
    ) -> list[RateRule]:
        rules = []
        for index, row in enumerate(rows, start=1):
            service = kilo_service if row["rate_type"] == "KILO" else pallet_service
            rules.append(
                RateRule(
                    rate_card=card,
                    service=service,
                    from_zone=row["from_zone"] or spec.origin_zone,
                    to_zone=row["to_zone"],
                    weight_min_kg=Decimal("0"),
                    weight_max_kg=None if row["rate_type"] == "KILO" else row["upto_kg"],
                    basic_charge=row["basic_charge"],
                    per_kg=row["per_kg"],
                    minimum_charge=row["minimum_charge"],
                    maximum_charge=None,
                    rule_type=RateRule.RuleType.LINEHAUL,
                    priority=index,
                    raw_payload={
                        "source": SOURCE_SYSTEM,
                        "source_sheet": spec.sheet_name,
                        "origin_zone": spec.origin_zone,
                        "rate_type": row["rate_type"],
                        "break_value": str(row["break_value"]),
                        "fuel_flag": row["fuel_flag"],
                        "upto_kg": str(row["upto_kg"]),
                        "cubic_factor": str(row["cubic_factor"]),
                        "source_rate": {
                            "basic_charge": str(row["basic_charge"]),
                            "per_kg": str(row["per_kg"]),
                            "minimum_charge": str(row["minimum_charge"]),
                        },
                        "source_row": row["source_row"],
                    },
                )
            )
        return rules

    def _zones(self, card: RateCard, spec: DfeSpec, rows: list[dict[str, Any]], allowed_zones: set[str]) -> list[RateZone]:
        return [
            RateZone(
                rate_card=card,
                origin_zone=spec.origin_zone,
                dest_zone=row["carrier_zone"],
                state=row["state"],
                suburb=row["suburb"],
                postcode=row["postcode"],
                deliverable=True,
                raw_payload={
                    "source": SOURCE_SYSTEM,
                    "source_zone_file": (card.metadata_json or {}).get("source_zone_file", ""),
                    "postcode_id": row["postcode_id"],
                    "carrier_zone": row["carrier_zone"],
                    "drop_code": row["drop_code"],
                    "sort_code": row["sort_code"],
                    "source_row": row["source_row"],
                },
            )
            for row in rows
            if row["carrier_zone"] in allowed_zones
        ]

    def _surcharge_rules(self, card: RateCard, destination_surcharges: list[dict[str, Any]]) -> list[SurchargeRule]:
        rules = [
            SurchargeRule(
                carrier=card.carrier,
                rate_card=card,
                code="FS",
                rule_name="DFE fuel levy",
                ratio=Decimal("0.196"),
                fee_amount=Decimal("0"),
                match_dimension=SurchargeRule.MatchDimension.ALWAYS,
                priority=1,
                active=True,
                raw_payload={
                    "source": SOURCE_SYSTEM,
                    "source_sheet": "Rate schedule conditions",
                    "note": "Direct Freight Express standard fuel levy 19.60%; configurable in Pricing -> Surcharges.",
                },
            )
        ]
        for index, row in enumerate(destination_surcharges, start=10):
            rules.append(
                SurchargeRule(
                    carrier=card.carrier,
                    rate_card=card,
                    code="DFE_DEST",
                    rule_name="DFE destination surcharge",
                    ratio=None,
                    fee_amount=row["amount"],
                    match_dimension=SurchargeRule.MatchDimension.ALWAYS,
                    condition_json={"postcode": row["postcode"], "suburb": row["suburb"]},
                    priority=index,
                    active=True,
                    raw_payload={
                        "source": SOURCE_SYSTEM,
                        "source_sheet": "Surcharge ",
                        "source_table": "DESTINATION SURCHARGE",
                        "fuel_levy_apply": True,
                    },
                )
            )
        return rules

    def _quote_channel(
        self,
        carrier: Carrier,
        service: CarrierService,
        card: RateCard,
        spec: DfeSpec,
        *,
        enabled: bool,
        pallet: bool,
    ) -> QuoteChannel:
        channel_code = spec.pallet_channel_code if pallet else spec.kilo_channel_code
        channel_name = spec.pallet_channel_name if pallet else spec.kilo_channel_name
        channel, _ = QuoteChannel.objects.update_or_create(
            code=channel_code,
            defaults={
                "name": channel_name,
                "carrier": carrier,
                "service": service,
                "provider_type": QuoteChannel.ProviderType.TABLE,
                "calculator_key": CALCULATOR_KEY,
                "quote_source": SOURCE_SYSTEM,
                "enabled": enabled,
                "priority": spec.priority + (100 if pallet else 0),
                "rate_card": card,
                "config_json": {
                    "source": SOURCE_SYSTEM,
                    "origin_zone": spec.origin_zone,
                    "rate_type": "PALLET" if pallet else "KILO",
                    "pallet_rows_imported_but_disabled": pallet,
                },
                "valid_from": spec.effective_from,
                "valid_to": None,
            },
        )
        return channel

    def _configure_defaults(self, platform_code: str, warehouse_code: str) -> dict[str, Any]:
        platform = Platform.objects.filter(code=platform_code, active=True).first() or Platform.objects.filter(active=True).order_by("id").first()
        warehouse = Warehouse.objects.filter(code=warehouse_code, active=True).first() or Warehouse.objects.filter(active=True).order_by("id").first()
        if not platform or not warehouse:
            return {"configured": False, "reason": "missing_active_platform_or_warehouse"}

        WarehousePlatform.objects.update_or_create(
            warehouse=warehouse,
            platform=platform,
            defaults={"enabled": True, "priority": 10, "is_default": True},
        )
        origin = self._warehouse_origin(warehouse)
        configured = []
        for spec in SPECS:
            if origin and spec.origin_zone != origin:
                continue
            carrier = self._find_imported_carrier()
            service = CarrierService.objects.filter(carrier=carrier, code=spec.kilo_service_code).first()
            if not carrier or not service:
                continue
            PlatformCarrier.objects.update_or_create(
                platform=platform,
                carrier=carrier,
                service=service,
                defaults={"enabled": True, "priority": spec.priority, "quote_source": SOURCE_SYSTEM},
            )
            WarehouseCarrier.objects.update_or_create(
                warehouse=warehouse,
                carrier=carrier,
                service=service,
                defaults={"enabled": True, "origin_zone": spec.origin_zone},
            )
            configured.append(f"{carrier.code}:{service.code}")
        return {
            "configured": bool(configured),
            "platform": platform.code,
            "warehouse": warehouse.code,
            "warehouse_origin": origin,
            "services": configured,
        }

    def _find_imported_carrier(self) -> Carrier | None:
        return (
            Carrier.objects.filter(source_payload_json__has_key="dfe_rate_proposal").order_by("id").first()
            or Carrier.objects.filter(name__icontains="Direct Freight").order_by("id").first()
            or Carrier.objects.filter(name__icontains="Direct").order_by("id").first()
        )

    def _warehouse_origin(self, warehouse: Warehouse) -> str:
        text = f"{warehouse.default_origin_zone} {warehouse.code} {warehouse.name} {warehouse.state} {warehouse.region}".upper()
        if "SYD" in text or "NSW" in text:
            return "SYDN"
        if "MEL" in text or "VIC" in text or warehouse.code.upper() == "BG01":
            return "MELB"
        return ""

    def _print_report(self, report: dict[str, Any], *, dry_run: bool) -> None:
        prefix = "DFE dry-run" if dry_run else "DFE import completed"
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix}: rate_cards={report.get('rate_cards', len(SPECS))}, "
                f"rate_rows={report.get('rate_rows_by_spec')}, zone_rows={report.get('zone_rows')}, "
                f"destination_surcharges={report.get('destination_surcharge_rows')}, "
                f"missing_zone_mappings={report.get('missing_zone_mappings')}"
            )
        )
        if report.get("default_configuration"):
            self.stdout.write(f"default_configuration={report['default_configuration']}")
