from __future__ import annotations

import json
import re
from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from freight.models import RateCard, RateRule, RateZone
from freight.quote_engine import json_safe


DEFAULT_RATE_FILE = (
    r"C:\Users\KenHu\.vscode\CourieDelivery\outputs\broers_rate_analysis\Broers\hunter"
    r"\Rate Card -  Hunter Road Freight - BroersGroupPtyLtd  - 20240920.xlsx"
)
DEFAULT_ZONE_FILE = (
    r"C:\Users\KenHu\.vscode\CourieDelivery\outputs\broers_rate_analysis\Broers\hunter"
    r"\Zone List - ZoneListCommon_01082025000320.xlsx"
)
DEFAULT_CARD_VERSION = "SP-HUNTER-SYD-2025"
SOURCE_SYSTEM = "BroersRatePackage"
SOURCE_LABEL = "Broers Hunter SYD 20240920"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_upper(value: Any) -> str:
    return clean(value).upper()


def clean_postcode(value: Any) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+(\.0+)?", text):
        text = text.split(".", 1)[0]
    return text.zfill(4) if text.isdigit() and len(text) < 4 else text


def dec(value: Any, default: str = "0") -> Decimal:
    text = clean(value).replace("$", "").replace(",", "")
    if not text or text.upper() in {"POA", "P.O.A", "P.O.A."}:
        text = default
    try:
        return Decimal(text)
    except (InvalidOperation, AttributeError):
        return Decimal(default)


def short_zone(display_zone: str, *parts: Any) -> str:
    base = re.sub(r"[^A-Z0-9]+", "_", clean_upper(display_zone or "RATE")).strip("_")[:24] or "RATE"
    joined = "|".join(clean(part) for part in parts)
    import hashlib

    digest = hashlib.sha1(joined.encode("utf-8")).hexdigest()[:8].upper()
    return f"{base}_{digest}"[:40]


class Command(BaseCommand):
    help = "Apply Broers Hunter Sydney rate workbook over the existing Hunter Sydney card without creating a new rate card."

    def add_arguments(self, parser):
        parser.add_argument("--rate-file", default=DEFAULT_RATE_FILE)
        parser.add_argument("--zone-file", default=DEFAULT_ZONE_FILE)
        parser.add_argument("--card-version", default=DEFAULT_CARD_VERSION)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--backup-dir", default=r"C:\Users\KenHu\.vscode\CourieDelivery\outputs\broers_rate_analysis")

    def handle(self, *args, **options):
        rate_file = Path(options["rate_file"])
        zone_file = Path(options["zone_file"])
        if not rate_file.exists():
            raise CommandError(f"Rate file not found: {rate_file}")
        if not zone_file.exists():
            raise CommandError(f"Zone file not found: {zone_file}")

        try:
            card = RateCard.objects.select_related("carrier", "service").get(version=options["card_version"])
        except RateCard.DoesNotExist as exc:
            raise CommandError(f"RateCard version not found: {options['card_version']}") from exc

        rates = self._load_sydney_rates(rate_file)
        zone_lookup = self._load_zone_lookup(zone_file)
        report = self._build_replacement(card, rates, zone_lookup)
        report["rate_file"] = str(rate_file)
        report["zone_file"] = str(zone_file)
        report["card_version"] = card.version
        report["card_name_before"] = card.name

        backup_path = self._write_backup(card, Path(options["backup_dir"]))
        report["backup_path"] = str(backup_path)

        if options["dry_run"]:
            public_report = {k: v for k, v in report.items() if not k.startswith("_")}
            self.stdout.write(json.dumps(public_report, ensure_ascii=False, indent=2))
            return

        with transaction.atomic():
            card.rules.all().delete()
            RateRule.objects.bulk_create(report.pop("_rules"), batch_size=1000)
            RateZone.objects.bulk_update(report.pop("_zones"), ["dest_zone", "raw_payload", "origin_zone"], batch_size=1000)

            meta = dict(card.metadata_json or {})
            meta["broers_rate_package_verification"] = {
                "source_package": "Broers rate",
                "source_file": str(rate_file),
                "zone_file": str(zone_file),
                "verification_date": timezone.now().isoformat(),
                "comparison": "Broers Hunter SYD differs from previous Hunter Sydney 2025; user confirmed HUNTER SYD should be source of truth.",
                "rate_match": False,
                "matched_rows": report["matched_rows"],
                "different_rows": report["updated_zone_rows"],
                "action": "Existing Hunter Sydney rate card overwritten with Broers Hunter SYD rates; no new rate card created.",
                "backup_path": str(backup_path),
            }
            meta["source"] = SOURCE_SYSTEM
            meta["source_rate_file"] = str(rate_file)
            meta["source_zone_file"] = str(zone_file)
            meta["source_effective_from"] = "2024-09-20"
            card.metadata_json = meta
            card.name = "Hunter SYD Broers 20240920"
            card.version_label = "Hunter SYD Broers 20240920"
            card.legacy_source_object = f"{SOURCE_SYSTEM}:Hunter_SYD_20240920"
            card.save(update_fields=["metadata_json", "name", "version_label", "legacy_source_object", "updated_at"])

            channel = card.quote_channels.first()
            if channel:
                channel.name = "Hunter SYD Broers 20240920"
                channel.quote_source = SOURCE_SYSTEM
                channel.save(update_fields=["name", "quote_source", "updated_at"])

        public_report = {k: v for k, v in report.items() if not k.startswith("_")}
        self.stdout.write(json.dumps(public_report, ensure_ascii=False, indent=2))

    def _load_sydney_rates(self, path: Path) -> dict[str, dict[str, Any]]:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [clean(value) for value in rows[0]]
        rates: dict[str, dict[str, Any]] = {}
        for values in rows[1:]:
            row = dict(zip(headers, values, strict=False))
            if clean_upper(row.get("From Zone")) != "SYDNEY":
                continue
            zone_number = clean(row.get("To Zone number"))
            if not zone_number:
                continue
            rates[zone_number] = {
                "from_zone": "SYD",
                "source_from_zone": clean(row.get("From Zone")),
                "source_from_zone_number": clean(row.get("From Zone number")),
                "source_to_zone": clean(row.get("To Zone")),
                "source_to_zone_number": zone_number,
                "rate_card_type": clean(row.get("Rate Card Type")),
                "minimum_charge": dec(row.get("Minimum Charge")),
                "basic": dec(row.get("Basic")),
                "per_kg": dec(row.get("Per KG")),
                "source_row": json_safe(row),
            }
        if not rates:
            raise CommandError("No SYDNEY rows found in Broers Hunter rate workbook.")
        return rates

    def _load_zone_lookup(self, path: Path) -> dict[tuple[str, str, str], dict[str, str]]:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [clean(value) for value in rows[0]]
        lookup: dict[tuple[str, str, str], dict[str, str]] = {}
        for values in rows[1:]:
            row = dict(zip(headers, values, strict=False))
            key = (clean_upper(row.get("State")), clean_upper(row.get("Suburb")), clean_postcode(row.get("Postcode")))
            customer_zone = clean(row.get("CustomerZone(Pricing Zones (Rate Region))"))
            if not all(key) or not customer_zone:
                continue
            lookup[key] = {
                "customer_zone": customer_zone,
                "label_zone": clean(row.get("LabelZone(Label/Sort Zones (Port) 2.0)")),
                "label_sub_zone": clean(row.get("LabelSubZone")),
                "source_row": json_safe(row),
            }
        return lookup

    def _build_replacement(
        self,
        card: RateCard,
        rates: dict[str, dict[str, Any]],
        zone_lookup: dict[tuple[str, str, str], dict[str, str]],
    ) -> dict[str, Any]:
        zones_to_update: list[RateZone] = []
        rules_by_key: OrderedDict[tuple[str, str, str, str], RateRule] = OrderedDict()
        missing_zone_map = []
        missing_rate = []

        for zone in card.zones.all().iterator(chunk_size=2000):
            key = (clean_upper(zone.state), clean_upper(zone.suburb), clean_postcode(zone.postcode))
            mapped = zone_lookup.get(key)
            if not mapped:
                missing_zone_map.append({"state": zone.state, "suburb": zone.suburb, "postcode": zone.postcode})
                continue
            rate = rates.get(mapped["customer_zone"])
            if not rate:
                missing_rate.append(
                    {
                        "state": zone.state,
                        "suburb": zone.suburb,
                        "postcode": zone.postcode,
                        "customer_zone": mapped["customer_zone"],
                    }
                )
                continue

            zone_code = short_zone(
                "BROERS_SYD",
                rate["from_zone"],
                rate["source_to_zone_number"],
                rate["basic"],
                rate["per_kg"],
                rate["minimum_charge"],
                zone.state,
            )
            rule_key = (zone_code, str(rate["basic"]), str(rate["per_kg"]), str(rate["minimum_charge"]))
            if rule_key not in rules_by_key:
                rules_by_key[rule_key] = RateRule(
                    rate_card=card,
                    service=card.service,
                    from_zone=rate["from_zone"],
                    to_zone=zone_code,
                    state="",
                    suburb="",
                    postcode="",
                    weight_min_kg=Decimal("0"),
                    weight_max_kg=None,
                    basic_charge=rate["basic"],
                    per_kg=rate["per_kg"],
                    minimum_charge=rate["minimum_charge"],
                    maximum_charge=None,
                    rule_type=RateRule.RuleType.LINEHAUL,
                    priority=len(rules_by_key) + 1,
                    raw_payload={
                        "source": SOURCE_SYSTEM,
                        "source_label": SOURCE_LABEL,
                        "source_rate": {
                            "basic": str(rate["basic"]),
                            "per_kg": str(rate["per_kg"]),
                            "minimum": str(rate["minimum_charge"]),
                        },
                        "source_to_zone": rate["source_to_zone"],
                        "source_to_zone_number": rate["source_to_zone_number"],
                        "source_row": rate["source_row"],
                    },
                )
            zone.origin_zone = rate["from_zone"]
            zone.dest_zone = zone_code
            raw = dict(zone.raw_payload or {})
            raw["source"] = SOURCE_SYSTEM
            raw["source_label"] = SOURCE_LABEL
            raw["display_zone"] = "BROERS_SYD"
            raw["hunter_customer_zone"] = mapped["customer_zone"]
            raw["hunter_label_zone"] = mapped["label_zone"]
            raw["hunter_label_sub_zone"] = mapped["label_sub_zone"]
            raw["source_rate"] = {
                "basic": str(rate["basic"]),
                "per_kg": str(rate["per_kg"]),
                "minimum": str(rate["minimum_charge"]),
                "source_to_zone": rate["source_to_zone"],
                "source_to_zone_number": rate["source_to_zone_number"],
            }
            raw["source_zone_row"] = mapped["source_row"]
            zone.raw_payload = raw
            zones_to_update.append(zone)

        if missing_zone_map or missing_rate:
            raise CommandError(
                "Cannot apply Broers Hunter SYD rates because some existing zones could not be mapped. "
                f"missing_zone_map={len(missing_zone_map)}, missing_rate={len(missing_rate)}"
            )

        return {
            "matched_rows": len(zones_to_update),
            "updated_zone_rows": len(zones_to_update),
            "new_rule_rows": len(rules_by_key),
            "missing_zone_map_count": len(missing_zone_map),
            "missing_rate_count": len(missing_rate),
            "_zones": zones_to_update,
            "_rules": list(rules_by_key.values()),
        }

    def _write_backup(self, card: RateCard, backup_dir: Path) -> Path:
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        path = backup_dir / f"hunter_sydney_before_broers_apply_{stamp}.json"
        payload = {
            "card": {
                "id": card.id,
                "name": card.name,
                "version": card.version,
                "metadata_json": card.metadata_json,
            },
            "rules": [
                {
                    "from_zone": rule.from_zone,
                    "to_zone": rule.to_zone,
                    "basic_charge": str(rule.basic_charge),
                    "per_kg": str(rule.per_kg),
                    "minimum_charge": str(rule.minimum_charge),
                    "raw_payload": rule.raw_payload,
                }
                for rule in card.rules.all().order_by("id")
            ],
            "zones": [
                {
                    "state": zone.state,
                    "suburb": zone.suburb,
                    "postcode": zone.postcode,
                    "origin_zone": zone.origin_zone,
                    "dest_zone": zone.dest_zone,
                    "raw_payload": zone.raw_payload,
                }
                for zone in card.zones.all().order_by("id")
            ],
        }
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return path
