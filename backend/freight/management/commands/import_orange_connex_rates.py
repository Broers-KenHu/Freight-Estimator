from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from freight.models import (
    Carrier,
    CarrierService,
    ImportJob,
    Platform,
    PlatformCarrier,
    QuoteChannel,
    RateCard,
    RateRule,
    RateZone,
    SurchargeRule,
    Warehouse,
    WarehouseCarrier,
    WarehousePlatform,
)

SOURCE_SYSTEM = "OrangeConnexEfn2026"
DEFAULT_RATE_FILE = r"C:\Users\KenHu\Downloads\2026 AU eFN rate.xlsx"
CALCULATOR_KEY = "freight.calculators.orange_connex_efn_2026.OrangeConnexEfn2026Calculator"
WEIGHT_BANDS = (
    ("Up to 250g", 0, 250),
    ("Up to 500g", 251, 500),
    ("501g-1kg", 501, 1000),
    ("1.01kg-2kg", 1001, 2000),
    ("2.01kg-3kg", 2001, 3000),
    ("3.01kg-4kg", 3001, 4000),
    ("4.01kg-5kg", 4001, 5000),
    ("5.01kg-7kg", 5001, 7000),
    ("7.01kg-10kg", 7001, 10000),
    ("10.01kg-15kg", 10001, 15000),
    ("15.01kg-22kg", 15001, 22000),
    ("22.01kg-25kg", 22001, 25000),
)


@dataclass(frozen=True)
class OrangeSpec:
    key: str
    origin_zone: str
    sheet_name: str
    rate_card_name: str
    version: str
    service_code: str
    service_name: str
    channel_code: str
    channel_name: str
    priority: int


SPECS = (
    OrangeSpec(
        key="orange_efn_mel_2026",
        origin_zone="MEL",
        sheet_name="PFL MEL",
        rate_card_name="Orange Connex eFN MEL 2026",
        version="ORANGE-EFN-MEL-2026",
        service_code="ORANGE_EFN_MEL_2026",
        service_name="Orange Connex eFN MEL 2026",
        channel_code="orange_efn_mel_2026",
        channel_name="Orange Connex eFN MEL 2026",
        priority=80,
    ),
    OrangeSpec(
        key="orange_efn_syd_2026",
        origin_zone="SYD",
        sheet_name="PFL SYD",
        rate_card_name="Orange Connex eFN SYD 2026",
        version="ORANGE-EFN-SYD-2026",
        service_code="ORANGE_EFN_SYD_2026",
        service_name="Orange Connex eFN SYD 2026",
        channel_code="orange_efn_syd_2026",
        channel_name="Orange Connex eFN SYD 2026",
        priority=81,
    ),
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_upper(value: Any) -> str:
    return clean(value).upper()


def clean_postcode(value: Any) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+(\.0+)?", text):
        text = text.split(".", 1)[0]
    return text.zfill(4) if text.isdigit() and len(text) < 4 else text


def dec(value: Any, default: str = "0") -> Decimal:
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        text = default
    try:
        return Decimal(text)
    except (InvalidOperation, AttributeError):
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return Decimal(match.group(0)) if match else Decimal(default)


def kg_from_grams(grams: int) -> Decimal:
    return Decimal(grams) / Decimal("1000")


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class Command(BaseCommand):
    help = "Import Orange Connex eFN 2026 fixed weight-band rate table."

    def add_arguments(self, parser):
        parser.add_argument("--rate-file", default=DEFAULT_RATE_FILE)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--configure-defaults", action="store_true")
        parser.add_argument("--carrier-name", default="Orange Connex")
        parser.add_argument("--default-platform-code", default="PI2022080502320043121506")
        parser.add_argument("--default-warehouse-code", default="BG01")

    def handle(self, *args, **options):
        rate_path = Path(options["rate_file"])
        if not rate_path.exists():
            raise CommandError(f"Orange Connex rate file not found: {rate_path}")
        workbook = load_workbook(rate_path, data_only=True, read_only=True)
        parsed_rates = {spec.key: self._parse_rate_sheet(workbook, spec) for spec in SPECS}
        zone_rows = self._parse_zone_rows(workbook)
        notes = self._parse_notes(workbook)
        report = self._source_report(rate_path, parsed_rates, zone_rows, notes)
        if options["dry_run"]:
            self._print_report(report, dry_run=True)
            return

        job = ImportJob.objects.create(
            job_type=ImportJob.JobType.RATE_CARD,
            status=ImportJob.Status.RUNNING,
            total_rows=report["total_import_rows"],
            report_json={"source": SOURCE_SYSTEM, "rate_file": str(rate_path)},
        )
        try:
            with transaction.atomic():
                import_report = self._import(options, parsed_rates, zone_rows, notes, report)
                if options["configure_defaults"]:
                    import_report["default_configuration"] = self._configure_defaults(
                        options["default_platform_code"],
                        options["default_warehouse_code"],
                    )
        except Exception as exc:  # noqa: BLE001
            job.status = ImportJob.Status.FAILED
            job.error_rows = report["total_import_rows"]
            job.progress = 100
            job.report_json = {**job.report_json, "error": str(exc)}
            job.save(update_fields=["status", "error_rows", "progress", "report_json", "updated_at"])
            raise

        job.status = ImportJob.Status.COMPLETED
        job.success_rows = report["total_import_rows"]
        job.error_rows = 0
        job.progress = 100
        job.report_json = {**job.report_json, **import_report}
        job.save(update_fields=["status", "success_rows", "error_rows", "progress", "report_json", "updated_at"])
        self._print_report(import_report, dry_run=False)

    def _parse_rate_sheet(self, workbook, spec: OrangeSpec) -> list[dict[str, Any]]:
        if spec.sheet_name not in workbook.sheetnames:
            raise CommandError(f"Missing Orange Connex rate sheet: {spec.sheet_name}")
        ws = workbook[spec.sheet_name]
        header = [clean(cell.value) for cell in next(ws.iter_rows(min_row=5, max_row=5))]
        band_columns = []
        for label, min_grams, max_grams in WEIGHT_BANDS:
            if label not in header:
                raise CommandError(f"Missing Orange Connex weight band {label} in {spec.sheet_name}")
            band_columns.append((header.index(label) + 1, label, min_grams, max_grams))

        rows = []
        for excel_row in ws.iter_rows(min_row=6, values_only=True):
            zone = clean(excel_row[1] if len(excel_row) > 1 else "")
            if not zone:
                break
            prices = [excel_row[col - 1] for col, *_ in band_columns]
            if not all(isinstance(price, (int, float, Decimal)) for price in prices):
                break
            for col, label, min_grams, max_grams in band_columns:
                rows.append(
                    {
                        "dest_zone": zone,
                        "band_label": label,
                        "min_grams": min_grams,
                        "max_grams": max_grams,
                        "price": dec(excel_row[col - 1]),
                    }
                )
        return rows

    def _parse_zone_rows(self, workbook) -> list[dict[str, Any]]:
        if "PFL Zone" not in workbook.sheetnames:
            raise CommandError("Missing Orange Connex PFL Zone sheet.")
        ws = workbook["PFL Zone"]
        rows = []
        seen: set[tuple[str, str, str, str]] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            suburb = clean_upper(row[0] if len(row) > 0 else "")
            postcode = clean_postcode(row[1] if len(row) > 1 else "")
            state = clean_upper(row[2] if len(row) > 2 else "")
            zone = clean(row[3] if len(row) > 3 else "")
            if not suburb or not postcode or not state or not zone:
                continue
            key = (suburb, postcode, state, zone)
            if key in seen:
                continue
            seen.add(key)
            rows.append({"suburb": suburb, "postcode": postcode, "state": state, "zone": zone})
        return rows

    def _parse_notes(self, workbook) -> dict[str, Any]:
        notes: dict[str, Any] = {
            "max_weight_kg": "25",
            "max_length_cm": "105",
            "max_volume_m3": "0.088",
            "tax_mode": "EX_GST",
            "valid_until": "2026-12-31",
            "other_charges": [],
        }
        ws = workbook["PFL MEL"]
        for row in ws.iter_rows(min_row=42, max_row=53, values_only=True):
            label = clean(row[1] if len(row) > 1 else "")
            value = clean(row[3] if len(row) > 3 else "") or clean(row[2] if len(row) > 2 else "")
            if label and value:
                notes["other_charges"].append({"label": label, "value": value})
        return notes

    def _source_report(
        self,
        rate_path: Path,
        parsed_rates: dict[str, list[dict[str, Any]]],
        zone_rows: list[dict[str, Any]],
        notes: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "source": SOURCE_SYSTEM,
            "rate_file": str(rate_path),
            "rate_file_sha256": file_sha256(rate_path),
            "rate_rules_by_spec": {key: len(rows) for key, rows in parsed_rates.items()},
            "rate_zones_by_spec": {key: len({row["dest_zone"] for row in rows}) for key, rows in parsed_rates.items()},
            "zone_rows": len(zone_rows),
            "zone_values": sorted({row["zone"] for row in zone_rows}),
            "notes": notes,
            "total_import_rows": sum(len(rows) for rows in parsed_rates.values()) + len(zone_rows) * len(SPECS),
        }

    def _import(
        self,
        options: dict[str, Any],
        parsed_rates: dict[str, list[dict[str, Any]]],
        zone_rows: list[dict[str, Any]],
        notes: dict[str, Any],
        source_report: dict[str, Any],
    ) -> dict[str, Any]:
        carrier = self._carrier(options["carrier_name"])
        report = {
            **source_report,
            "carrier_id": carrier.id,
            "carrier_code": carrier.code,
            "carrier_name": carrier.name,
            "rate_cards": 0,
            "rate_zones": 0,
            "rate_rules": 0,
            "surcharge_rules": 0,
            "quote_channels": 0,
        }
        for spec in SPECS:
            service = self._service(carrier, spec)
            card = self._rate_card(carrier, service, spec, parsed_rates[spec.key], notes, source_report)
            card.rules.all().delete()
            card.zones.all().delete()
            card.surcharge_rules.all().delete()
            rules = self._rate_rules(card, service, spec, parsed_rates[spec.key])
            zones = self._zones(card, zone_rows)
            surcharges = self._reference_surcharges(card, notes)
            RateRule.objects.bulk_create(rules, batch_size=1000)
            RateZone.objects.bulk_create(zones, batch_size=1000)
            SurchargeRule.objects.bulk_create(surcharges, batch_size=100)
            self._quote_channel(carrier, service, card, spec)
            report["rate_cards"] += 1
            report["rate_zones"] += len(zones)
            report["rate_rules"] += len(rules)
            report["surcharge_rules"] += len(surcharges)
            report["quote_channels"] += 1
        return report

    def _carrier(self, carrier_name: str) -> Carrier:
        carrier = Carrier.objects.filter(name__icontains="Orange Connex").order_by("id").first()
        if not carrier:
            carrier = Carrier.objects.filter(name__icontains="orange").order_by("id").first()
        if not carrier:
            carrier = Carrier.objects.create(
                code=self._next_carrier_code(),
                name=carrier_name,
                carrier_type=Carrier.CarrierType.TABLE,
                active=True,
                source_system=SOURCE_SYSTEM,
            )
        original_name = carrier.name
        update_fields = []
        if carrier.name != carrier_name:
            carrier.name = carrier_name
            update_fields.append("name")
        if carrier.carrier_type == Carrier.CarrierType.API:
            carrier.carrier_type = Carrier.CarrierType.HYBRID
            update_fields.append("carrier_type")
        payload = dict(carrier.source_payload_json or {})
        payload["orange_connex_efn_2026"] = {
            "source": SOURCE_SYSTEM,
            "last_imported_at": timezone.now().isoformat(),
            "previous_carrier_name": original_name,
        }
        carrier.source_payload_json = payload
        update_fields.append("source_payload_json")
        carrier.save(update_fields=[*set(update_fields), "updated_at"])
        return carrier

    def _next_carrier_code(self) -> str:
        max_number = 0
        for code in Carrier.objects.filter(code__startswith="CAR").values_list("code", flat=True):
            match = re.fullmatch(r"CAR(\d{6})", code or "")
            if match:
                max_number = max(max_number, int(match.group(1)))
        next_number = max_number + 1
        while True:
            code = f"CAR{next_number:06d}"
            if not Carrier.objects.filter(code=code).exists():
                return code
            next_number += 1

    def _service(self, carrier: Carrier, spec: OrangeSpec) -> CarrierService:
        service, _ = CarrierService.objects.update_or_create(
            carrier=carrier,
            code=spec.service_code,
            defaults={"name": spec.service_name, "service_level": spec.origin_zone, "active": True},
        )
        return service

    def _rate_card(
        self,
        carrier: Carrier,
        service: CarrierService,
        spec: OrangeSpec,
        rows: list[dict[str, Any]],
        notes: dict[str, Any],
        source_report: dict[str, Any],
    ) -> RateCard:
        metadata = {
            "source": SOURCE_SYSTEM,
            "source_rate_file": source_report["rate_file"],
            "source_rate_file_sha256": source_report["rate_file_sha256"],
            "source_sheet": spec.sheet_name,
            "origin_zone": spec.origin_zone,
            "calculator_key": CALCULATOR_KEY,
            "pricing_mode": "ARTICLE_WEIGHT_BAND",
            "weight_bands": [
                {"label": label, "min_grams": min_grams, "max_grams": max_grams}
                for label, min_grams, max_grams in WEIGHT_BANDS
            ],
            "notes": notes,
            "rate_rows": len(rows),
            "destination_zones": sorted({row["dest_zone"] for row in rows}),
        }
        card, _ = RateCard.objects.update_or_create(
            legacy_source_object=f"{SOURCE_SYSTEM}:{spec.version}",
            version=spec.version,
            defaults={
                "carrier": carrier,
                "service": service,
                "name": spec.rate_card_name,
                "version_label": spec.version,
                "status": RateCard.Status.ACTIVE,
                "effective_from": date(2026, 1, 1),
                "effective_to": date(2026, 12, 31),
                "is_active": True,
                "priority": spec.priority,
                "currency": "AUD",
                "tax_mode": RateCard.TaxMode.EX_GST,
                "gst_rate": Decimal("0.10"),
                "cubic_factor": Decimal("0"),
                "metadata_json": metadata,
            },
        )
        return card

    def _rate_rules(
        self,
        card: RateCard,
        service: CarrierService,
        spec: OrangeSpec,
        rows: list[dict[str, Any]],
    ) -> list[RateRule]:
        rules = []
        for index, row in enumerate(rows, start=1):
            rules.append(
                RateRule(
                    rate_card=card,
                    service=service,
                    from_zone=spec.origin_zone,
                    to_zone=row["dest_zone"],
                    weight_min_kg=kg_from_grams(row["min_grams"]),
                    weight_max_kg=kg_from_grams(row["max_grams"]),
                    basic_charge=row["price"],
                    per_kg=Decimal("0"),
                    minimum_charge=Decimal("0"),
                    rule_type=RateRule.RuleType.LINEHAUL,
                    priority=index,
                    raw_payload={
                        "source": SOURCE_SYSTEM,
                        "source_sheet": spec.sheet_name,
                        "origin_zone": spec.origin_zone,
                        "band_label": row["band_label"],
                        "min_grams": row["min_grams"],
                        "max_grams": row["max_grams"],
                        "pricing_mode": "ARTICLE_WEIGHT_BAND",
                    },
                )
            )
        return rules

    def _zones(self, card: RateCard, rows: list[dict[str, Any]]) -> list[RateZone]:
        return [
            RateZone(
                rate_card=card,
                origin_zone=(card.metadata_json or {}).get("origin_zone", ""),
                dest_zone=row["zone"],
                state=row["state"],
                suburb=row["suburb"],
                postcode=row["postcode"],
                deliverable=True,
                raw_payload={"source": SOURCE_SYSTEM, "source_sheet": "PFL Zone"},
            )
            for row in rows
        ]

    def _reference_surcharges(self, card: RateCard, notes: dict[str, Any]) -> list[SurchargeRule]:
        rules = []
        for index, charge in enumerate(notes.get("other_charges") or [], start=1):
            amount = dec(charge.get("value"))
            if amount <= 0:
                continue
            rules.append(
                SurchargeRule(
                    carrier=card.carrier,
                    rate_card=card,
                    code=self._charge_code(charge["label"]),
                    rule_name=charge["label"],
                    fee_amount=amount,
                    match_dimension=SurchargeRule.MatchDimension.ALWAYS,
                    priority=100 + index,
                    active=False,
                    raw_payload={
                        "source": SOURCE_SYSTEM,
                        "note": "Reference only. Not applied automatically without operational event data.",
                        "source_value": charge["value"],
                    },
                )
            )
        return rules

    def _charge_code(self, label: str) -> str:
        base = re.sub(r"[^A-Z0-9]+", "_", label.upper()).strip("_")
        return f"ORANGE_{base}"[:40]

    def _quote_channel(self, carrier: Carrier, service: CarrierService, card: RateCard, spec: OrangeSpec) -> QuoteChannel:
        channel, _ = QuoteChannel.objects.update_or_create(
            code=spec.channel_code,
            defaults={
                "name": spec.channel_name,
                "carrier": carrier,
                "service": service,
                "provider_type": QuoteChannel.ProviderType.TABLE,
                "calculator_key": CALCULATOR_KEY,
                "quote_source": SOURCE_SYSTEM,
                "enabled": True,
                "priority": spec.priority,
                "rate_card": card,
                "config_json": {"source": SOURCE_SYSTEM, "origin_zone": spec.origin_zone},
                "valid_from": date(2026, 1, 1),
                "valid_to": date(2026, 12, 31),
            },
        )
        return channel

    def _configure_defaults(self, platform_code: str, warehouse_code: str) -> dict[str, Any]:
        platform = Platform.objects.filter(code=platform_code, active=True).first() or Platform.objects.filter(active=True).order_by("id").first()
        warehouse = Warehouse.objects.filter(code=warehouse_code, active=True).first() or Warehouse.objects.filter(active=True).order_by("id").first()
        if not platform or not warehouse:
            return {"configured": False, "reason": "missing_active_platform_or_warehouse"}
        WarehousePlatform.objects.update_or_create(
            warehouse=warehouse,
            platform=platform,
            defaults={"enabled": True, "priority": 10, "is_default": True},
        )
        origin = self._warehouse_origin(warehouse)
        carrier = Carrier.objects.filter(name="Orange Connex").order_by("id").first()
        configured = []
        for spec in SPECS:
            if origin and spec.origin_zone != origin:
                continue
            service = CarrierService.objects.filter(carrier=carrier, code=spec.service_code).first()
            if not carrier or not service:
                continue
            PlatformCarrier.objects.update_or_create(
                platform=platform,
                carrier=carrier,
                service=service,
                defaults={"enabled": True, "priority": spec.priority, "quote_source": SOURCE_SYSTEM},
            )
            WarehouseCarrier.objects.update_or_create(
                warehouse=warehouse,
                carrier=carrier,
                service=service,
                defaults={"enabled": True, "origin_zone": spec.origin_zone},
            )
            configured.append(f"{carrier.code}:{service.code}")
        return {"configured": bool(configured), "platform": platform.code, "warehouse": warehouse.code, "warehouse_origin": origin, "services": configured}

    def _warehouse_origin(self, warehouse: Warehouse) -> str:
        text = f"{warehouse.default_origin_zone} {warehouse.code} {warehouse.name} {warehouse.state} {warehouse.region}".upper()
        if "SYD" in text or "NSW" in text:
            return "SYD"
        if "MEL" in text or "VIC" in text or warehouse.code.upper() == "BG01":
            return "MEL"
        return ""

    def _print_report(self, report: dict[str, Any], *, dry_run: bool) -> None:
        prefix = "Orange Connex dry-run" if dry_run else "Orange Connex import completed"
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix}: rate_cards={report.get('rate_cards', len(SPECS))}, "
                f"rate_rules={report.get('rate_rules_by_spec')}, zone_rows={report.get('zone_rows')}, "
                f"zone_values={len(report.get('zone_values', []))}"
            )
        )
        if report.get("default_configuration"):
            self.stdout.write(f"default_configuration={report['default_configuration']}")
