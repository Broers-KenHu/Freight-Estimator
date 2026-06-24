from __future__ import annotations

import csv
import datetime as dt
import os
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

import psycopg
import pymssql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from psycopg.rows import dict_row


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = PROJECT_ROOT / "backend" / ".env"
REPORT_DIR = PROJECT_ROOT / "reports"


@dataclass
class LocalRow:
    tracking: str
    local_order_key: str
    local_erp_order_no: str
    local_erp_owner_order_no: str
    local_rd3_order_id: str
    local_platform_order_no: str
    local_invoice_no: str
    local_actual_freight: Any
    local_match_status: str
    local_invoice_source: str
    local_invoice_carrier: str
    local_invoice_service: str
    local_erp_carrier: str
    local_erp_channel: str
    local_erp_service: str
    local_reconciliation_item_id: Any
    local_erp_snapshot_id: Any
    local_invoice_snapshot_id: Any


def load_env(path: Path) -> dict[str, str]:
    env = dict(os.environ)
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def clean(value: Any) -> str:
    return str(value or "").strip()


def norm(value: Any) -> str:
    return clean(value).upper()


def connect_local(env: dict[str, str]):
    return psycopg.connect(env["DATABASE_URL"], connect_timeout=20, row_factory=dict_row)


def connect_invoice_reader(env: dict[str, str]):
    return pymssql.connect(
        server=env.get("INVOICE_SQLSERVER_HOST") or env.get("SQLSERVER_HOST") or "192.168.72.8",
        port=int(env.get("INVOICE_SQLSERVER_PORT") or env.get("SQLSERVER_PORT") or "1433"),
        user=env.get("INVOICE_SQLSERVER_USER") or env.get("SQLSERVER_USER"),
        password=env.get("INVOICE_SQLSERVER_PASSWORD") or env.get("SQLSERVER_PASSWORD"),
        database=env.get("INVOICE_SQLSERVER_DATABASE") or "invoiceReader",
        login_timeout=20,
        timeout=60,
        as_dict=True,
    )


def load_local_by_tracking(conn) -> dict[str, list[LocalRow]]:
    query = """
        SELECT
            COALESCE(NULLIF(ics.tracking_no, ''), NULLIF(ess.tracking_no, ''), NULLIF(iri.consignment_no, '')) AS tracking,
            COALESCE(NULLIF(ess.erp_owner_order_no, ''), NULLIF(ess.erp_order_no, ''), NULLIF(iri.order_no, '')) AS local_order_key,
            COALESCE(NULLIF(ess.erp_order_no, ''), NULLIF(iri.order_no, '')) AS local_erp_order_no,
            COALESCE(NULLIF(ess.erp_owner_order_no, ''), '') AS local_erp_owner_order_no,
            COALESCE(NULLIF(ess.third_party_order_no, ''), '') AS local_rd3_order_id,
            COALESCE(NULLIF(ess.platform_order_no, ''), '') AS local_platform_order_no,
            COALESCE(NULLIF(ics.invoice_no, ''), '') AS local_invoice_no,
            iri.actual_freight AS local_actual_freight,
            iri.match_status AS local_match_status,
            COALESCE(NULLIF(ics.source_label, ''), NULLIF(ics.source_key, ''), '') AS local_invoice_source,
            COALESCE(NULLIF(ics.carrier_name, ''), '') AS local_invoice_carrier,
            COALESCE(NULLIF(ics.service_name, ''), '') AS local_invoice_service,
            COALESCE(NULLIF(ess.carrier_name, ''), '') AS local_erp_carrier,
            COALESCE(NULLIF(ess.carrier_channel, ''), '') AS local_erp_channel,
            COALESCE(NULLIF(ess.service_provider, ''), '') AS local_erp_service,
            iri.id AS local_reconciliation_item_id,
            ess.id AS local_erp_snapshot_id,
            ics.id AS local_invoice_snapshot_id
        FROM invoice_reconciliation_item iri
        LEFT JOIN invoice_charge_snapshot ics ON ics.id = iri.invoice_charge_snapshot_id
        LEFT JOIN erp_shipment_snapshot ess ON ess.id = iri.erp_shipment_snapshot_id
        WHERE (iri.source_system LIKE 'invoiceReader%%' OR ics.id IS NOT NULL)
          AND COALESCE(NULLIF(ics.tracking_no, ''), NULLIF(ess.tracking_no, ''), NULLIF(iri.consignment_no, '')) IS NOT NULL
    """
    result: dict[str, list[LocalRow]] = defaultdict(list)
    with conn.cursor(name="local_mismatch_sample_cursor") as cur:
        cur.itersize = 20000
        cur.execute(query)
        while True:
            rows = cur.fetchmany(20000)
            if not rows:
                break
            for row in rows:
                tracking = norm(row["tracking"])
                local_order_key = norm(row["local_order_key"])
                if not tracking or not local_order_key:
                    continue
                result[tracking].append(
                    LocalRow(
                        tracking=tracking,
                        local_order_key=local_order_key,
                        local_erp_order_no=clean(row["local_erp_order_no"]),
                        local_erp_owner_order_no=clean(row["local_erp_owner_order_no"]),
                        local_rd3_order_id=clean(row["local_rd3_order_id"]),
                        local_platform_order_no=clean(row["local_platform_order_no"]),
                        local_invoice_no=clean(row["local_invoice_no"]),
                        local_actual_freight=row["local_actual_freight"],
                        local_match_status=clean(row["local_match_status"]),
                        local_invoice_source=clean(row["local_invoice_source"]),
                        local_invoice_carrier=clean(row["local_invoice_carrier"]),
                        local_invoice_service=clean(row["local_invoice_service"]),
                        local_erp_carrier=clean(row["local_erp_carrier"]),
                        local_erp_channel=clean(row["local_erp_channel"]),
                        local_erp_service=clean(row["local_erp_service"]),
                        local_reconciliation_item_id=row["local_reconciliation_item_id"],
                        local_erp_snapshot_id=row["local_erp_snapshot_id"],
                        local_invoice_snapshot_id=row["local_invoice_snapshot_id"],
                    )
                )
    return result


def sql_optional_columns(conn) -> set[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT c.name
            FROM sys.columns c
            WHERE c.object_id = OBJECT_ID('dbo.erp_match_results')
            """
        )
        return {str(row["name"]) for row in cur.fetchall()}


def sql_select(column: str, available: set[str], alias: str | None = None) -> str:
    alias = alias or column
    if column in available:
        return f"[{column}] AS [{alias}]"
    return f"CAST(NULL AS NVARCHAR(500)) AS [{alias}]"


def export_samples(
    sql_conn,
    local_by_tracking: dict[str, list[LocalRow]],
    limit: int = 100,
    *,
    require_local_erp_snapshot: bool = True,
) -> list[dict[str, Any]]:
    available = sql_optional_columns(sql_conn)
    optional = [
        sql_select("detail_source_key", available),
        sql_select("detail_source_label", available),
        sql_select("detail_table", available),
        sql_select("detail_carrier", available),
        sql_select("detail_service", available),
        sql_select("carrier_name", available, "invoice_reader_carrier_name"),
        sql_select("service_name", available, "invoice_reader_service_name"),
        sql_select("match_method", available),
        sql_select("match_confidence", available),
        sql_select("match_reason", available),
    ]
    query = f"""
        SELECT
            CONVERT(NVARCHAR(500), [detail_tracking]) AS [detail_tracking],
            CONVERT(NVARCHAR(500), [erp_owner_order_no]) AS [erp_owner_order_no],
            CONVERT(NVARCHAR(500), [erp_rd3_order_id]) AS [erp_rd3_order_id],
            CONVERT(NVARCHAR(500), [invoice_no]) AS [invoice_no],
            TRY_CONVERT(DECIMAL(18,4), [detail_amount_inc_gst]) AS [detail_amount_inc_gst],
            {", ".join(optional)}
        FROM [dbo].[erp_match_results]
        WHERE NULLIF(LTRIM(RTRIM(CONVERT(NVARCHAR(500), [detail_tracking]))), '') IS NOT NULL
          AND NULLIF(LTRIM(RTRIM(CONVERT(NVARCHAR(500), [erp_owner_order_no]))), '') IS NOT NULL
        ORDER BY [detail_tracking], [invoice_no]
    """
    samples: list[dict[str, Any]] = []
    seen_tracking: set[str] = set()
    with sql_conn.cursor() as cur:
        cur.execute(query)
        while len(samples) < limit:
            rows = cur.fetchmany(20000)
            if not rows:
                break
            for sql_row in rows:
                tracking = norm(sql_row.get("detail_tracking"))
                if not tracking or tracking in seen_tracking:
                    continue
                sql_order = norm(sql_row.get("erp_owner_order_no"))
                if not sql_order:
                    continue
                locals_for_tracking = local_by_tracking.get(tracking)
                if not locals_for_tracking:
                    continue
                if require_local_erp_snapshot:
                    locals_for_tracking = [row for row in locals_for_tracking if row.local_erp_snapshot_id]
                    if not locals_for_tracking:
                        continue
                local_orders = {row.local_order_key for row in locals_for_tracking if row.local_order_key}
                if not local_orders or sql_order in local_orders:
                    continue
                local = locals_for_tracking[0]
                seen_tracking.add(tracking)
                samples.append(
                    {
                        "tracking": clean(sql_row.get("detail_tracking")),
                        "invoice_reader_erp_owner_order_no": clean(sql_row.get("erp_owner_order_no")),
                        "couriedelivery_local_order_key": " | ".join(sorted(local_orders)),
                        "invoice_reader_erp_rd3_order_id": clean(sql_row.get("erp_rd3_order_id")),
                        "local_erp_order_no": local.local_erp_order_no,
                        "local_erp_owner_order_no": local.local_erp_owner_order_no,
                        "local_rd3_order_id": local.local_rd3_order_id,
                        "local_platform_order_no": local.local_platform_order_no,
                        "invoice_reader_invoice_no": clean(sql_row.get("invoice_no")),
                        "local_invoice_no": local.local_invoice_no,
                        "invoice_reader_amount_inc_gst": sql_row.get("detail_amount_inc_gst"),
                        "local_actual_freight_inc_gst": local.local_actual_freight,
                        "local_match_status": local.local_match_status,
                        "invoice_reader_source_key": clean(sql_row.get("detail_source_key")),
                        "invoice_reader_source_label": clean(sql_row.get("detail_source_label")),
                        "invoice_reader_detail_table": clean(sql_row.get("detail_table")),
                        "invoice_reader_detail_carrier": clean(sql_row.get("detail_carrier") or sql_row.get("invoice_reader_carrier_name")),
                        "invoice_reader_detail_service": clean(sql_row.get("detail_service") or sql_row.get("invoice_reader_service_name")),
                        "invoice_reader_match_method": clean(sql_row.get("match_method")),
                        "invoice_reader_match_confidence": clean(sql_row.get("match_confidence")),
                        "invoice_reader_match_reason": clean(sql_row.get("match_reason")),
                        "local_invoice_source": local.local_invoice_source,
                        "local_invoice_carrier": local.local_invoice_carrier,
                        "local_invoice_service": local.local_invoice_service,
                        "local_erp_carrier": local.local_erp_carrier,
                        "local_erp_channel": local.local_erp_channel,
                        "local_erp_service": local.local_erp_service,
                        "local_reconciliation_item_id": local.local_reconciliation_item_id,
                        "local_erp_snapshot_id": local.local_erp_snapshot_id,
                        "local_invoice_snapshot_id": local.local_invoice_snapshot_id,
                    }
                )
                if len(samples) >= limit:
                    break
    return samples


def with_basis(samples: list[dict[str, Any]], basis: str) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in samples:
        output.append({"sample_basis": basis, **row})
    return output


def recommended_samples(strict_samples: list[dict[str, Any]], all_samples: list[dict[str, Any]], limit: int = 100) -> list[dict[str, Any]]:
    strict_trackings = {norm(row.get("tracking")) for row in strict_samples}
    combined = with_basis(strict_samples[:limit], "STRICT_BOTH_ERP_MAPPED")
    if len(combined) >= limit:
        return combined[:limit]
    fill_rows = [row for row in all_samples if norm(row.get("tracking")) not in strict_trackings]
    combined.extend(with_basis(fill_rows[: limit - len(combined)], "BROADER_CONFLICT_FILLER"))
    return combined[:limit]


def write_csv(path: Path, samples: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(samples[0].keys()))
        writer.writeheader()
        writer.writerows(samples)


def apply_sheet_format(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 22,
        "B": 26,
        "C": 34,
        "D": 26,
        "E": 24,
        "F": 26,
        "G": 26,
        "H": 26,
        "I": 22,
        "J": 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        if letter not in widths:
            sheet.column_dimensions[letter].width = 20
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_rows(sheet, samples: list[dict[str, Any]]) -> None:
    headers = list(samples[0].keys())
    sheet.append(headers)
    for row in samples:
        sheet.append([row.get(header) for header in headers])
    apply_sheet_format(sheet)


def write_xlsx(
    path: Path,
    recommended: list[dict[str, Any]],
    strict_samples: list[dict[str, Any]],
    all_samples: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "recommended_100"
    append_rows(sheet, recommended)

    strict_sheet = workbook.create_sheet("both_erp_mapped")
    append_rows(strict_sheet, with_basis(strict_samples, "STRICT_BOTH_ERP_MAPPED"))

    all_sheet = workbook.create_sheet("all_conflicts")
    append_rows(all_sheet, with_basis(all_samples, "BROADER_CONFLICT"))

    note_sheet = workbook.create_sheet("notes")
    note_sheet.append(["sheet", "description"])
    note_sheet.append(["recommended_100", "100 rows for review. Strict ERP-mapped mismatches are listed first; filler rows are broader conflicts marked in sample_basis."])
    note_sheet.append(["both_erp_mapped", "Strict samples where InvoiceReader has erp_owner_order_no and CourieDelivery has local ErpShipmentSnapshot, but order keys differ. Use these first for ERP validation."])
    note_sheet.append(["all_conflicts", "100 broader same-tracking different-order samples, including local UNMATCHED rows where CourieDelivery only has invoice/order references."])
    apply_sheet_format(note_sheet)
    workbook.save(path)


def main() -> None:
    env = load_env(ENV_PATH)
    REPORT_DIR.mkdir(exist_ok=True)
    today = dt.date.today().strftime("%Y%m%d")
    csv_path = REPORT_DIR / f"invoice_reader_same_tracking_different_order_samples_{today}.csv"
    xlsx_path = REPORT_DIR / f"invoice_reader_same_tracking_different_order_samples_{today}.xlsx"

    with connect_local(env) as local_conn:
        local_by_tracking = load_local_by_tracking(local_conn)
    with connect_invoice_reader(env) as sql_conn:
        strict_samples = export_samples(sql_conn, local_by_tracking, limit=100, require_local_erp_snapshot=True)
        all_samples = export_samples(sql_conn, local_by_tracking, limit=100, require_local_erp_snapshot=False)

    if not strict_samples:
        raise RuntimeError("No same-tracking different-order samples were found.")

    recommended = recommended_samples(strict_samples, all_samples, limit=100)
    write_csv(csv_path, recommended)
    write_xlsx(xlsx_path, recommended, strict_samples, all_samples)
    print(csv_path)
    print(xlsx_path)
    print(f"recommended_samples={len(recommended)}")
    print(f"strict_samples={len(strict_samples)}")
    print(f"all_conflicts_samples={len(all_samples)}")


if __name__ == "__main__":
    main()
